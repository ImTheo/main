#!/usr/bin/env python3
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import sys
import io
from pathlib import Path

# Importar la lógica original
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string
import re

# ── constantes del script original ──────────────────────────────────────────
EXCEL_MAX_COLUMN = 16384
EXCEL_COLUMN_RE = re.compile(r"^[A-Za-z]{1,3}$")
NOT_FOUND_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")

# ── lógica original (copiada para no depender de args) ───────────────────────
def normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower()

def normalize_match_value(value):
    if value is None:
        return ""
    text = str(value).replace('"', "").strip()
    if not text:
        return ""
    if text.endswith(".0"):
        integer_candidate = text[:-2]
        if integer_candidate.lstrip("-").isdigit():
            text = integer_candidate
    words = text.split()
    if not words:
        return ""
    return words[0].lower()

def is_column_letter(value):
    if not EXCEL_COLUMN_RE.match(value):
        return False
    try:
        column_index = column_index_from_string(value.upper())
    except ValueError:
        return False
    return column_index <= EXCEL_MAX_COLUMN

def resolve_column(ws, header_row, column_spec, workbook_label):
    spec = str(column_spec).strip()
    if not spec:
        raise ValueError(f"Columna no encontrada en archivo {workbook_label}: {column_spec}")
    if spec.isdigit():
        column_index = int(spec)
        if column_index <= 0 or column_index > ws.max_column:
            raise ValueError(f"Columna inválida en archivo {workbook_label}: {column_spec}")
        return column_index
    if is_column_letter(spec):
        column_index = column_index_from_string(spec.upper())
        if column_index > ws.max_column:
            raise ValueError(f"Columna no encontrada en archivo {workbook_label}: {column_spec}")
        return column_index
    wanted_header = normalize_header(spec)
    for cell in ws[header_row]:
        if normalize_header(cell.value) == wanted_header:
            return cell.column
    raise ValueError(f"Columna no encontrada en archivo {workbook_label}: '{column_spec}'")

def combine_lookup_values(values):
    if len(values) == 1:
        return values[0]
    return ", ".join("" if v is None else str(v) for v in values)

def build_lookup_map(ws, header_row, match_col, value_col):
    lookup_values = {}
    seen_keys = set()
    duplicated_keys = set()
    duplicated_count = 0
    for row in range(header_row + 1, ws.max_row + 1):
        key = normalize_match_value(ws.cell(row=row, column=match_col).value)
        if not key:
            continue
        if key in seen_keys:
            duplicated_count += 1
            duplicated_keys.add(key)
        seen_keys.add(key)
        lookup_values.setdefault(key, []).append(ws.cell(row=row, column=value_col).value)
    lookup = {k: combine_lookup_values(v) for k, v in lookup_values.items()}
    return lookup, duplicated_count, sorted(duplicated_keys)

def mark_row_yellow(ws, row):
    for column in range(1, ws.max_column + 1):
        ws.cell(row=row, column=column).fill = NOT_FOUND_FILL

def run_process(dynamic_file, lookup_file, dynamic_header_row, lookup_header_row,
                dynamic_match_col, dynamic_replace_col, lookup_match_col, lookup_value_col):
    """Ejecuta la lógica principal y retorna un dict con los resultados."""
    dynamic_wb = load_workbook(dynamic_file)
    lookup_wb  = load_workbook(lookup_file)
    dynamic_ws = dynamic_wb.worksheets[0]
    lookup_ws  = lookup_wb.worksheets[0]

    dm_col = resolve_column(dynamic_ws, dynamic_header_row, dynamic_match_col,   "principal")
    dr_col = resolve_column(dynamic_ws, dynamic_header_row, dynamic_replace_col, "principal")
    lm_col = resolve_column(lookup_ws,  lookup_header_row,  lookup_match_col,    "referencia")
    lv_col = resolve_column(lookup_ws,  lookup_header_row,  lookup_value_col,    "referencia")

    lookup_map, duplicated_count, duplicated_keys = build_lookup_map(
        lookup_ws, lookup_header_row, lm_col, lv_col)

    processed_rows = 0
    updated_rows   = 0
    rows_without   = 0
    dynamic_keys   = set()

    for row in range(dynamic_header_row + 1, dynamic_ws.max_row + 1):
        key = normalize_match_value(dynamic_ws.cell(row=row, column=dm_col).value)
        if not key:
            continue
        dynamic_keys.add(key)
        processed_rows += 1
        if key not in lookup_map:
            rows_without += 1
            mark_row_yellow(dynamic_ws, row)
            continue
        dynamic_ws.cell(row=row, column=dr_col).value = lookup_map[key]
        updated_rows += 1

    output_file = Path(dynamic_file).with_name(f"{Path(dynamic_file).stem}_updated.xlsx")
    dynamic_wb.save(output_file)

    matching_dups = [k for k in duplicated_keys if k in dynamic_keys]

    return {
        "processed": processed_rows,
        "updated":   updated_rows,
        "no_match":  rows_without,
        "output":    str(output_file),
        "dup_count": duplicated_count,
        "dup_keys":  matching_dups,
    }

# ── GUI ───────────────────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Mapeador de Archivos Excel")
        self.resizable(True, True)
        self.minsize(720, 600)
        self.geometry("780x820")
        self.configure(bg="#F0F4F8")

        # Paleta
        BG      = "#F0F4F8"
        CARD    = "#FFFFFF"
        ACCENT  = "#2563EB"
        ACCENT2 = "#1D4ED8"
        TEXT    = "#1E293B"
        SUBTEXT = "#64748B"
        BORDER  = "#CBD5E1"
        SUCCESS = "#16A34A"
        WARN    = "#D97706"

        self.configure(bg=BG)

        # ── título ────────────────────────────────────────────────────────────
        header = tk.Frame(self, bg=ACCENT, pady=18)
        header.pack(fill="x")
        tk.Label(header, text="📊  Mapeador de Archivos Excel",
                 font=("Segoe UI", 16, "bold"), bg=ACCENT, fg="white").pack()
        tk.Label(header, text="Actualiza columnas de un archivo usando otro como referencia",
                 font=("Segoe UI", 9), bg=ACCENT, fg="#BFDBFE").pack()

        # ── contenedor principal con scroll ──────────────────────────────────
        canvas = tk.Canvas(self, bg=BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        main = tk.Frame(canvas, bg=BG, padx=28, pady=20)
        canvas_window = canvas.create_window((0, 0), window=main, anchor="nw")

        def on_frame_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        main.bind("<Configure>", on_frame_configure)

        def on_canvas_configure(e):
            canvas.itemconfig(canvas_window, width=e.width)
        canvas.bind("<Configure>", on_canvas_configure)

        def on_mousewheel(e):
            canvas.yview_scroll(int(-1*(e.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", on_mousewheel)

        def section(parent, title):
            f = tk.LabelFrame(parent, text=f"  {title}  ",
                              font=("Segoe UI", 10, "bold"),
                              bg=CARD, fg=ACCENT, bd=1, relief="solid",
                              padx=16, pady=12)
            f.pack(fill="x", pady=8)
            return f

        def file_row(parent, label_text, var, row):
            tk.Label(parent, text=label_text, font=("Segoe UI", 9, "bold"),
                     bg=CARD, fg=TEXT, anchor="w").grid(row=row, column=0, sticky="w", pady=4)
            e = tk.Entry(parent, textvariable=var, font=("Segoe UI", 9),
                         width=46, relief="solid", bd=1)
            e.grid(row=row, column=1, padx=8, pady=4)
            btn = tk.Button(parent, text="📂 Examinar",
                            font=("Segoe UI", 9), bg=ACCENT, fg="white",
                            relief="flat", cursor="hand2", padx=10,
                            command=lambda v=var: self._browse(v))
            btn.grid(row=row, column=2, pady=4)
            btn.bind("<Enter>", lambda e, b=btn: b.configure(bg=ACCENT2))
            btn.bind("<Leave>", lambda e, b=btn: b.configure(bg=ACCENT))

        def param_row(parent, label_text, var, row, col_offset=0, tooltip=None):
            r, c = row, col_offset
            lbl = tk.Label(parent, text=label_text, font=("Segoe UI", 9),
                           bg=CARD, fg=TEXT, anchor="w")
            lbl.grid(row=r, column=c, sticky="w", padx=(0,6), pady=5)
            e = tk.Entry(parent, textvariable=var, font=("Segoe UI", 9),
                         width=22, relief="solid", bd=1)
            e.grid(row=r, column=c+1, sticky="w", pady=5)
            if tooltip:
                tk.Label(parent, text=f"ℹ {tooltip}", font=("Segoe UI", 7),
                         bg=CARD, fg=SUBTEXT).grid(row=r+1, column=c, columnspan=2,
                                                    sticky="w", pady=(0,4))

        # Variables
        self.v_dynamic_file      = tk.StringVar()
        self.v_lookup_file       = tk.StringVar()
        self.v_dynamic_hrow      = tk.StringVar(value="5")
        self.v_lookup_hrow       = tk.StringVar(value="1")
        self.v_dynamic_match_col = tk.StringVar(value="Nº serie")
        self.v_dynamic_repl_col  = tk.StringVar(value="Nombre del equipo")
        self.v_lookup_match_col  = tk.StringVar(value="Nº serie")
        self.v_lookup_value_col  = tk.StringVar(value="Nombre")

        # ── Archivos ──────────────────────────────────────────────────────────
        sec_files = section(main, "📁  Archivos")
        sec_files.columnconfigure(1, weight=1)
        file_row(sec_files, "Archivo principal (.xlsx):", self.v_dynamic_file, 0)
        tk.Label(sec_files, text="El archivo que se va a actualizar",
                 font=("Segoe UI", 7), bg=CARD, fg=SUBTEXT
                 ).grid(row=1, column=1, sticky="w", padx=8)
        file_row(sec_files, "Archivo de referencia (.xlsx):", self.v_lookup_file, 2)
        tk.Label(sec_files, text="El archivo del que se toman los valores nuevos",
                 font=("Segoe UI", 7), bg=CARD, fg=SUBTEXT
                 ).grid(row=3, column=1, sticky="w", padx=8)

        # ── Parámetros ────────────────────────────────────────────────────────
        sec_params = section(main, "⚙️  Configuración de columnas")
        sec_params.columnconfigure((0,1,2,3), weight=1)

        tk.Label(sec_params, text="ARCHIVO PRINCIPAL", font=("Segoe UI", 8, "bold"),
                 bg=CARD, fg=ACCENT).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0,4))
        tk.Label(sec_params, text="ARCHIVO DE REFERENCIA", font=("Segoe UI", 8, "bold"),
                 bg=CARD, fg=ACCENT).grid(row=0, column=2, columnspan=2, sticky="w", pady=(0,4))

        tk.Label(sec_params, text="Fila de encabezado:", font=("Segoe UI", 9),
                 bg=CARD, fg=TEXT).grid(row=1, column=0, sticky="w", pady=4)
        tk.Entry(sec_params, textvariable=self.v_dynamic_hrow,
                 font=("Segoe UI", 9), width=6, relief="solid", bd=1
                 ).grid(row=1, column=1, sticky="w", pady=4)

        tk.Label(sec_params, text="Fila de encabezado:", font=("Segoe UI", 9),
                 bg=CARD, fg=TEXT).grid(row=1, column=2, sticky="w", padx=(20,0), pady=4)
        tk.Entry(sec_params, textvariable=self.v_lookup_hrow,
                 font=("Segoe UI", 9), width=6, relief="solid", bd=1
                 ).grid(row=1, column=3, sticky="w", pady=4)

        tk.Label(sec_params, text="Columna de coincidencia:", font=("Segoe UI", 9),
                 bg=CARD, fg=TEXT).grid(row=2, column=0, sticky="w", pady=4)
        tk.Entry(sec_params, textvariable=self.v_dynamic_match_col,
                 font=("Segoe UI", 9), width=20, relief="solid", bd=1
                 ).grid(row=2, column=1, sticky="w", pady=4)

        tk.Label(sec_params, text="Columna de coincidencia:", font=("Segoe UI", 9),
                 bg=CARD, fg=TEXT).grid(row=2, column=2, sticky="w", padx=(20,0), pady=4)
        tk.Entry(sec_params, textvariable=self.v_lookup_match_col,
                 font=("Segoe UI", 9), width=20, relief="solid", bd=1
                 ).grid(row=2, column=3, sticky="w", pady=4)

        tk.Label(sec_params, text="Columna a reemplazar:", font=("Segoe UI", 9),
                 bg=CARD, fg=TEXT).grid(row=3, column=0, sticky="w", pady=4)
        tk.Entry(sec_params, textvariable=self.v_dynamic_repl_col,
                 font=("Segoe UI", 9), width=20, relief="solid", bd=1
                 ).grid(row=3, column=1, sticky="w", pady=4)

        tk.Label(sec_params, text="Columna con valores nuevos:", font=("Segoe UI", 9),
                 bg=CARD, fg=TEXT).grid(row=3, column=2, sticky="w", padx=(20,0), pady=4)
        tk.Entry(sec_params, textvariable=self.v_lookup_value_col,
                 font=("Segoe UI", 9), width=20, relief="solid", bd=1
                 ).grid(row=3, column=3, sticky="w", pady=4)

        tk.Label(sec_params,
                 text="ℹ  Puedes usar el nombre exacto de la columna, la letra (A, B, C…) o el número",
                 font=("Segoe UI", 7), bg=CARD, fg=SUBTEXT
                 ).grid(row=4, column=0, columnspan=4, sticky="w", pady=(4,0))

        # ── Botón ejecutar ────────────────────────────────────────────────────
        btn_frame = tk.Frame(main, bg=BG)
        btn_frame.pack(pady=6)

        self.btn_run = tk.Button(btn_frame, text="▶  Ejecutar proceso",
                                 font=("Segoe UI", 11, "bold"),
                                 bg=SUCCESS, fg="white", relief="flat",
                                 cursor="hand2", padx=30, pady=10,
                                 command=self._start)
        self.btn_run.pack()
        self.btn_run.bind("<Enter>", lambda e: self.btn_run.configure(bg="#15803D"))
        self.btn_run.bind("<Leave>", lambda e: self.btn_run.configure(bg=SUCCESS))

        # ── Barra de progreso ─────────────────────────────────────────────────
        self.progress = ttk.Progressbar(main, mode="indeterminate", length=500)
        self.progress.pack(pady=(4,0))

        # ── Resultados ────────────────────────────────────────────────────────
        sec_res = section(main, "📋  Resultados")
        self.txt_result = tk.Text(sec_res, font=("Consolas", 9), height=12,
                                  bg="#F8FAFC", fg=TEXT, relief="solid", bd=1,
                                  state="disabled", wrap="word")
        self.txt_result.pack(fill="x")

        # ── pie ───────────────────────────────────────────────────────────────
        tk.Label(main, text="Las filas sin coincidencia quedan marcadas en amarillo en el archivo de salida.",
                 font=("Segoe UI", 8), bg=BG, fg=SUBTEXT).pack(pady=(0,10))

    # ── helpers ───────────────────────────────────────────────────────────────
    def _browse(self, var):
        path = filedialog.askopenfilename(
            filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")])
        if path:
            var.set(path)

    def _log(self, text, tag=None):
        self.txt_result.configure(state="normal")
        self.txt_result.insert("end", text)
        if tag:
            start = self.txt_result.index("end-1c linestart")
            self.txt_result.tag_add(tag, start, "end-1c")
        self.txt_result.see("end")
        self.txt_result.configure(state="disabled")

    def _clear_log(self):
        self.txt_result.configure(state="normal")
        self.txt_result.delete("1.0", "end")
        self.txt_result.configure(state="disabled")

    def _start(self):
        # Validaciones básicas
        if not self.v_dynamic_file.get():
            messagebox.showwarning("Archivo faltante", "Selecciona el archivo principal.")
            return
        if not self.v_lookup_file.get():
            messagebox.showwarning("Archivo faltante", "Selecciona el archivo de referencia.")
            return

        self.btn_run.configure(state="disabled", text="⏳  Procesando…")
        self.progress.start(10)
        self._clear_log()
        self._log("Iniciando proceso…\n")

        threading.Thread(target=self._run, daemon=True).start()

    def _run(self):
        try:
            result = run_process(
                dynamic_file       = self.v_dynamic_file.get(),
                lookup_file        = self.v_lookup_file.get(),
                dynamic_header_row = int(self.v_dynamic_hrow.get()),
                lookup_header_row  = int(self.v_lookup_hrow.get()),
                dynamic_match_col  = self.v_dynamic_match_col.get(),
                dynamic_replace_col= self.v_dynamic_repl_col.get(),
                lookup_match_col   = self.v_lookup_match_col.get(),
                lookup_value_col   = self.v_lookup_value_col.get(),
            )
            self.after(0, self._show_result, result)
        except Exception as ex:
            self.after(0, self._show_error, str(ex))

    def _show_result(self, r):
        self.progress.stop()
        self.btn_run.configure(state="normal", text="▶  Ejecutar proceso")
        self._clear_log()

        if r["dup_count"]:
            self._log(f"⚠  {r['dup_count']} claves duplicadas encontradas (valores unidos con comas).\n")
            if r["dup_keys"]:
                self._log("   Duplicados que coinciden con el archivo principal:\n")
                for k in r["dup_keys"]:
                    self._log(f"   • {k}\n")
            self._log("\n")

        self._log(f"✅  Proceso completado\n\n")
        self._log(f"   Filas procesadas:       {r['processed']}\n")
        self._log(f"   Filas actualizadas:     {r['updated']}\n")
        self._log(f"   Filas sin coincidencia: {r['no_match']}\n")
        self._log(f"\n📄  Archivo generado:\n   {r['output']}\n")

        messagebox.showinfo("¡Listo!",
            f"Proceso completado.\n\n"
            f"✅ Actualizadas: {r['updated']} filas\n"
            f"❌ Sin coincidencia: {r['no_match']} filas\n\n"
            f"Archivo guardado en:\n{r['output']}")

    def _show_error(self, msg):
        self.progress.stop()
        self.btn_run.configure(state="normal", text="▶  Ejecutar proceso")
        self._log(f"❌  Error: {msg}\n")
        messagebox.showerror("Error", msg)


if __name__ == "__main__":
    app = App()
    app.mainloop()
