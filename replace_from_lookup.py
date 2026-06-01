#!/usr/bin/env python3
import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string


EXCEL_MAX_COLUMN = 16384
EXCEL_COLUMN_RE = re.compile(r"^[A-Za-z]{1,3}$")
NOT_FOUND_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")


def fail(message):
    print(f"Error: {message}", file=sys.stderr)
    sys.exit(1)


def validate_xlsx_file(path):
    if path.suffix.lower() != ".xlsx":
        fail(f"Only .xlsx files are supported: {path}")
    if not path.exists():
        fail(f"File not found: {path}")
    if not path.is_file():
        fail(f"File not found: {path}")


def parse_header_row(value):
    try:
        number = int(value)
    except ValueError:
        fail(f"Invalid header row: {value}")

    if number <= 0:
        fail(f"Invalid header row: {value}")
    return number


def normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def normalize_match_value(value):
    if value is None:
        return ""

    text = str(value).replace('"', "").strip()
    if not text:
        return ""

    if text.endswith(".0"):
        integer_candidate = text[:-2]
        if integer_candidate.lstrip("-").isdigit():
            text = integer_candidate

    words = text.split()
    if not words:
        return ""

    return words[0].lower()


def is_column_letter(value):
    if not EXCEL_COLUMN_RE.match(value):
        return False

    try:
        column_index = column_index_from_string(value.upper())
    except ValueError:
        return False

    return column_index <= EXCEL_MAX_COLUMN


def resolve_column(ws, header_row, column_spec, workbook_label):
    spec = str(column_spec).strip()
    if not spec:
        fail(f"Column not found in {workbook_label} file: {column_spec}")

    if spec.isdigit():
        column_index = int(spec)
        validate_column_index(column_index, ws, column_spec, workbook_label)
        return column_index

    if is_column_letter(spec):
        try:
            column_index = column_index_from_string(spec.upper())
        except ValueError:
            fail(f"Invalid column in {workbook_label} file: {column_spec}")
        validate_column_index(column_index, ws, column_spec, workbook_label)
        return column_index

    wanted_header = normalize_header(spec)
    for cell in ws[header_row]:
        if normalize_header(cell.value) == wanted_header:
            return cell.column

    fail(f"Column not found in {workbook_label} file: {column_spec}")


def validate_column_index(column_index, ws, column_spec, workbook_label):
    if column_index <= 0 or column_index > EXCEL_MAX_COLUMN:
        fail(f"Invalid column in {workbook_label} file: {column_spec}")
    if column_index > ws.max_column:
        fail(f"Column not found in {workbook_label} file: {column_spec}")


def load_xlsx(path):
    try:
        return load_workbook(path)
    except Exception:
        fail(f"Could not open workbook: {path}")


def build_lookup_map(ws, header_row, match_col, value_col):
    lookup = {}
    seen_keys = set()
    duplicated_keys = set()
    duplicated_count = 0

    for row in range(header_row + 1, ws.max_row + 1):
        key = normalize_match_value(ws.cell(row=row, column=match_col).value)
        if not key:
            continue

        if key in seen_keys:
            duplicated_count += 1
            duplicated_keys.add(key)
        seen_keys.add(key)
        lookup[key] = ws.cell(row=row, column=value_col).value

    return lookup, duplicated_count, sorted(duplicated_keys)


def mark_row_yellow(ws, row):
    for column in range(1, ws.max_column + 1):
        ws.cell(row=row, column=column).fill = NOT_FOUND_FILL


def output_path_for(dynamic_file):
    return dynamic_file.with_name(f"{dynamic_file.stem}_updated.xlsx")


def parse_args():
    parser = argparse.ArgumentParser(
        description="Update values in one .xlsx file using another .xlsx file as a lookup matrix."
    )
    parser.add_argument("--dynamic-file", required=True, type=Path)
    parser.add_argument("--lookup-file", required=True, type=Path)
    parser.add_argument("--dynamic-header-row", required=True)
    parser.add_argument("--lookup-header-row", required=True)
    parser.add_argument("--dynamic-match-column", required=True)
    parser.add_argument("--dynamic-replace-column", required=True)
    parser.add_argument("--lookup-match-column", required=True)
    parser.add_argument("--lookup-value-column", required=True)
    parser.add_argument(
        "--print-duplicated-lookup-keys",
        action="store_true",
        help="Print normalized duplicated lookup keys that also match the dynamic file.",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    dynamic_header_row = parse_header_row(args.dynamic_header_row)
    lookup_header_row = parse_header_row(args.lookup_header_row)

    validate_xlsx_file(args.dynamic_file)
    validate_xlsx_file(args.lookup_file)

    dynamic_wb = load_xlsx(args.dynamic_file)
    lookup_wb = load_xlsx(args.lookup_file)

    dynamic_ws = dynamic_wb.worksheets[0]
    lookup_ws = lookup_wb.worksheets[0]

    if dynamic_header_row > dynamic_ws.max_row:
        fail(f"Invalid header row: {dynamic_header_row}")
    if lookup_header_row > lookup_ws.max_row:
        fail(f"Invalid header row: {lookup_header_row}")

    dynamic_match_col = resolve_column(
        dynamic_ws,
        dynamic_header_row,
        args.dynamic_match_column,
        "dynamic",
    )
    dynamic_replace_col = resolve_column(
        dynamic_ws,
        dynamic_header_row,
        args.dynamic_replace_column,
        "dynamic",
    )
    lookup_match_col = resolve_column(
        lookup_ws,
        lookup_header_row,
        args.lookup_match_column,
        "lookup",
    )
    lookup_value_col = resolve_column(
        lookup_ws,
        lookup_header_row,
        args.lookup_value_column,
        "lookup",
    )

    lookup_map, duplicated_count, duplicated_keys = build_lookup_map(
        lookup_ws,
        lookup_header_row,
        lookup_match_col,
        lookup_value_col,
    )

    processed_rows = 0
    updated_rows = 0
    rows_without_match = 0
    dynamic_keys = set()

    for row in range(dynamic_header_row + 1, dynamic_ws.max_row + 1):
        key = normalize_match_value(dynamic_ws.cell(row=row, column=dynamic_match_col).value)

        if not key:
            continue

        dynamic_keys.add(key)
        processed_rows += 1

        if key not in lookup_map:
            rows_without_match += 1
            mark_row_yellow(dynamic_ws, row)
            continue

        dynamic_ws.cell(row=row, column=dynamic_replace_col).value = lookup_map[key]
        updated_rows += 1

    output_file = output_path_for(args.dynamic_file)
    dynamic_wb.save(output_file)

    if duplicated_count:
        if args.print_duplicated_lookup_keys:
            matching_duplicated_keys = [
                key for key in duplicated_keys if key in dynamic_keys
            ]
            if matching_duplicated_keys:
                print("Duplicated lookup keys matching dynamic file:")
                for key in matching_duplicated_keys:
                    print(f"- {key}")
            else:
                print("No duplicated lookup keys match the dynamic file.")

    print(f"Processed rows: {processed_rows}")
    print(f"Updated rows: {updated_rows}")
    print(f"Rows without match: {rows_without_match}")
    print(f"Output file: {output_file}")


if __name__ == "__main__":
    main()
